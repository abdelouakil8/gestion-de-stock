"""PDF and XLSX generation for store reports.

Reuses the same Money/Decimal discipline and existing statistics queries.
"""

from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, numbers
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Store, StoreSettings
from app.services import alerts, customers, day_closing, sales, statistics

_A4_WIDTH, _A4_HEIGHT = A4
_MARGIN = 20 * mm
_LINE = 6 * mm

# Register Arabic font (same font file used by receipts.py)
_FONT_DIR = Path(__file__).resolve().parents[1] / "assets" / "fonts"
pdfmetrics.registerFont(TTFont("Amiri", str(_FONT_DIR / "Amiri-Regular.ttf")))

# Characters that make Excel/LibreOffice treat a cell as a formula.
_DANGEROUS_CELL_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

# Mapping for payment methods — kept for external callers; PDF/XLSX now use i18n.
PAYMENT_LABELS = {
    "cash": "Espèces",
    "card": "Carte",
    "mobile": "Mobile",
    "other": "Autre",
}

_REPORT_I18N: dict[str, dict[str, str]] = {
    "fr": {
        "title": "RAPPORT DE SYNTHÈSE",
        "generated": "Généré le",
        "date_range": "Du {f} au {t}",
        "sec_financial": "Résumé financier",
        "sec_top_products": "Meilleures ventes",
        "sec_top_customers": "Meilleurs clients",
        "sec_payments": "Répartition par mode de paiement",
        "sec_low_stock": "Stock faible",
        "sec_credits": "Crédits en attente",
        "kpi_revenue": "Chiffre d'affaires :",
        "kpi_gross_profit": "Marge brute :",
        "kpi_sales_count": "Nombre de ventes :",
        "kpi_discounts": "Total remises :",
        "col_product": "Produit",
        "col_quantity": "Quantité",
        "col_revenue": "CA",
        "col_customer": "Client",
        "col_phone": "Téléphone",
        "col_sales": "Ventes",
        "col_mode": "Mode",
        "col_amount": "Montant",
        "col_transactions": "Transactions",
        "col_stock": "Stock",
        "col_threshold": "Seuil",
        "col_total": "Total",
        "col_paid": "Payé",
        "col_balance": "Reste",
        "col_age": "Ancienneté",
        "unknown": "Inconnu",
        "day_abbr": "j",
        "sheet_summary": "Synthèse",
        "sheet_top_products": "Meilleures ventes",
        "sheet_top_customers": "Meilleurs clients",
        "sheet_payments": "Modes de paiement",
        "sheet_low_stock": "Stock faible",
        "sheet_credits": "Crédits en attente",
        "xlsx_metric": "Métrique",
        "xlsx_value": "Valeur",
        "xlsx_revenue": "Chiffre d'affaires",
        "xlsx_gross_profit": "Marge brute",
        "xlsx_num_sales": "Nombre de ventes",
        "xlsx_discounts": "Total remises",
        "xlsx_qty_sold": "Quantité vendue",
        "xlsx_current_stock": "Stock actuel",
        "xlsx_alert_threshold": "Seuil d'alerte",
        "xlsx_age_days": "Ancienneté (jours)",
        "payment_cash": "Espèces",
        "payment_card": "Carte",
        "payment_mobile": "Mobile",
        "payment_other": "Autre",
        "kpi_stock_cost": "Valeur du stock (coût) :",
        "kpi_stock_retail": "Valeur du stock (vente) :",
        "kpi_supplier_debt": "Dette fournisseurs :",
        "sec_category": "Ventes par catégorie",
        "sec_dead_stock": "Stock dormant",
        "col_category": "Catégorie",
        "col_profit": "Bénéfice",
        "col_margin": "Marge",
        "col_tied": "Capital figé",
        "col_days_dormant": "Jours sans vente",
        "no_category": "Sans catégorie",
        "never_sold": "Jamais",
        "sheet_category": "Catégories",
        "sheet_dead_stock": "Stock dormant",
        "xlsx_profit": "Bénéfice",
        "xlsx_quantity": "Quantité",
        "xlsx_stock_cost": "Valeur du stock (coût)",
        "xlsx_stock_retail": "Valeur du stock (vente)",
        "xlsx_supplier_debt": "Dette fournisseurs",
        "xlsx_tied": "Capital figé",
        "xlsx_days_dormant": "Jours sans vente",
        "debt_title": "CRÉANCES CLIENTS",
        "debt_total_label": "Total dû :",
        "debtors_label": "Débiteurs :",
        "daily_title": "RAPPORT JOURNALIER",
        "daily_sec_summary": "Résumé des ventes",
        "daily_sec_payments": "Par mode de paiement",
        "daily_sec_top": "Top 10 des produits",
        "daily_sec_log": "Journal détaillé des ventes",
        "kpi_transactions": "Nombre de transactions :",
        "kpi_net_revenue": "Chiffre d'affaires net :",
        "kpi_avg_basket": "Panier moyen :",
        "kpi_refunds": "Total remboursements :",
        "payment_transfer": "Virement",
        "col_time": "Heure",
        "col_invoice": "N° facture",
        "col_items": "Articles",
        "col_payment": "Paiement",
        "signature_manager": "Signature du responsable",
        "anonymous": "Anonyme",
        "comparison_title": "COMPARAISON DE PÉRIODES",
        "cmp_period_a": "Période A",
        "cmp_period_b": "Période B",
        "cmp_change": "Évolution",
        "cmp_basket": "Panier moyen",
    },
    "ar": {
        "title": "تقرير الملخص",
        "generated": "تم الإنشاء",
        "date_range": "من {f} إلى {t}",
        "sec_financial": "الملخص المالي",
        "sec_top_products": "أفضل المبيعات",
        "sec_top_customers": "أفضل العملاء",
        "sec_payments": "توزيع وسائل الدفع",
        "sec_low_stock": "مخزون منخفض",
        "sec_credits": "قروض معلقة",
        "kpi_revenue": ": رقم المبيعات",
        "kpi_gross_profit": ": الهامش الإجمالي",
        "kpi_sales_count": ": عدد المبيعات",
        "kpi_discounts": ": مجموع الخصومات",
        "col_product": "المنتج",
        "col_quantity": "الكمية",
        "col_revenue": "رقم المبيعات",
        "col_customer": "العميل",
        "col_phone": "الهاتف",
        "col_sales": "المبيعات",
        "col_mode": "الوسيلة",
        "col_amount": "المبلغ",
        "col_transactions": "المعاملات",
        "col_stock": "المخزون",
        "col_threshold": "العتبة",
        "col_total": "الإجمالي",
        "col_paid": "المدفوع",
        "col_balance": "المتبقي",
        "col_age": "قدم الدين",
        "unknown": "مجهول",
        "day_abbr": "ي",
        "sheet_summary": "الملخص",
        "sheet_top_products": "أفضل المبيعات",
        "sheet_top_customers": "أفضل العملاء",
        "sheet_payments": "وسائل الدفع",
        "sheet_low_stock": "مخزون منخفض",
        "sheet_credits": "قروض معلقة",
        "xlsx_metric": "المقياس",
        "xlsx_value": "القيمة",
        "xlsx_revenue": "رقم المبيعات",
        "xlsx_gross_profit": "الهامش الإجمالي",
        "xlsx_num_sales": "عدد المبيعات",
        "xlsx_discounts": "مجموع الخصومات",
        "xlsx_qty_sold": "الكمية المباعة",
        "xlsx_current_stock": "المخزون الحالي",
        "xlsx_alert_threshold": "عتبة التنبيه",
        "xlsx_age_days": "قدم الدين (أيام)",
        "payment_cash": "نقدا",
        "payment_card": "بطاقة",
        "payment_mobile": "جوال",
        "payment_other": "أخرى",
        "kpi_stock_cost": ": قيمة المخزون (التكلفة)",
        "kpi_stock_retail": ": قيمة المخزون (البيع)",
        "kpi_supplier_debt": ": ديون الموردين",
        "sec_category": "المبيعات حسب الفئة",
        "sec_dead_stock": "المخزون الراكد",
        "col_category": "الفئة",
        "col_profit": "الربح",
        "col_margin": "الهامش",
        "col_tied": "رأس المال المجمّد",
        "col_days_dormant": "أيام بلا بيع",
        "no_category": "بدون فئة",
        "never_sold": "أبداً",
        "sheet_category": "الفئات",
        "sheet_dead_stock": "المخزون الراكد",
        "xlsx_profit": "الربح",
        "xlsx_quantity": "الكمية",
        "xlsx_stock_cost": "قيمة المخزون (التكلفة)",
        "xlsx_stock_retail": "قيمة المخزون (البيع)",
        "xlsx_supplier_debt": "ديون الموردين",
        "xlsx_tied": "رأس المال المجمّد",
        "xlsx_days_dormant": "أيام بلا بيع",
        "debt_title": "ديون العملاء",
        "debt_total_label": ": إجمالي المستحق",
        "debtors_label": ": المدينون",
        "daily_title": "التقرير اليومي",
        "daily_sec_summary": "ملخص المبيعات",
        "daily_sec_payments": "حسب وسيلة الدفع",
        "daily_sec_top": "أفضل 10 منتجات",
        "daily_sec_log": "سجل المبيعات المفصل",
        "kpi_transactions": ": عدد المعاملات",
        "kpi_net_revenue": ": صافي رقم المبيعات",
        "kpi_avg_basket": ": متوسط السلة",
        "kpi_refunds": ": إجمالي المبالغ المرجعة",
        "payment_transfer": "تحويل",
        "col_time": "الوقت",
        "col_invoice": "رقم الفاتورة",
        "col_items": "الأصناف",
        "col_payment": "الدفع",
        "signature_manager": "توقيع المسؤول",
        "anonymous": "مجهول",
        "comparison_title": "مقارنة الفترات",
        "cmp_period_a": "الفترة أ",
        "cmp_period_b": "الفترة ب",
        "cmp_change": "التغير",
        "cmp_basket": "متوسط السلة",
    },
}


def _sanitize_cell(value):
    """Neutralize CSV/Excel formula injection.

    A spreadsheet treats a text cell starting with ``=``, ``+``, ``-``, ``@``
    (or a leading tab/carriage-return) as a formula. Any user-controlled
    string written to a report — product/customer/category names, phone
    numbers, arbitrary payment-method codes — is forced to literal text by
    prefixing a single quote. Non-strings pass through untouched.
    """
    if isinstance(value, str) and value.startswith(_DANGEROUS_CELL_PREFIXES):
        return "'" + value
    return value


def _shape_arabic(text: str) -> str:
    """Shape and reorder Arabic text for ReportLab rendering."""
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        configuration = {
            "delete_harakat": False,
            "shift_harakat_position": False,
            "use_unshaped_instead_of_isolated": False,
        }
        reshaper = arabic_reshaper.ArabicReshaper(configuration=configuration)
        reshaped = reshaper.reshape(text)
        return get_display(reshaped)
    except ImportError:
        return text.encode("latin-1", "replace").decode("latin-1")


def _safe(text: str, is_arabic: bool = False) -> str:
    """Degrade Latin text or shape Arabic text for ReportLab."""
    if is_arabic:
        return _shape_arabic(text)
    return text.encode("latin-1", "replace").decode("latin-1")


def _fmt_money(value: Decimal | int | str | None) -> str:
    """French monetary format: '1 234 567,50 DA' (NBSP thousands, comma dec)."""
    amount = Decimal(str(value if value is not None else 0)).quantize(Decimal("0.01"))
    text = f"{amount:,.2f}".replace(",", " ").replace(".", ",")
    return f"{text} DA"


def build_summary_report_pdf(
    db: Session,
    store_id: UUID,
    date_from: datetime,
    date_to: datetime,
    language: str = "fr",
) -> bytes:
    """A4 PDF report spanning sales, products, customers, stock, and credits."""
    store = db.scalar(select(Store).where(Store.id == store_id))
    store_name = store.name if store else "Ma Boutique"

    is_arabic = language == "ar"
    i18n = _REPORT_I18N.get(language, _REPORT_I18N["fr"])
    default_font = "Amiri" if is_arabic else "Helvetica"
    bold_font = "Amiri" if is_arabic else "Helvetica-Bold"

    payment_labels = {
        "cash": i18n["payment_cash"],
        "card": i18n["payment_card"],
        "mobile": i18n["payment_mobile"],
        "other": i18n["payment_other"],
    }

    # Fetch data
    summary = statistics.sales_summary(db, store_id, date_from, date_to)
    top_prods = statistics.top_products(db, store_id, date_from, date_to, limit=10)
    top_custs = customers.top_customers(db, store_id, date_from, date_to, limit=10)
    category_stats = statistics.category_breakdown(db, store_id, date_from, date_to)
    payments = statistics.payment_method_breakdown(db, store_id, date_from, date_to)
    inventory = statistics.inventory_stats(db, store_id)
    financial = statistics.financial_snapshot(db, store_id)
    low_stock = alerts.low_stock_products(db, store_id)
    dead = statistics.dead_stock(db, store_id, days=60, limit=15)
    credits = alerts.outstanding_credits(db, store_id)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    y = _A4_HEIGHT - _MARGIN
    page_num = 1

    def safe(text: str) -> str:
        return _safe(str(text), is_arabic)

    def draw_str(x: float, y_pos: float, text: str) -> None:
        """LTR left-anchor (fr) or RTL right-anchor (ar), mirroring x."""
        if is_arabic:
            pdf.drawRightString(_A4_WIDTH - x, y_pos, text)
        else:
            pdf.drawString(x, y_pos, text)

    def new_page_if_needed(needed: float) -> None:
        nonlocal y, page_num
        if y - needed < _MARGIN:
            _draw_footer()
            pdf.showPage()
            y = _A4_HEIGHT - _MARGIN
            page_num += 1
            _draw_header()

    def _draw_header() -> None:
        nonlocal y
        pdf.setFont(bold_font, 18)
        pdf.drawCentredString(_A4_WIDTH / 2, y, safe(i18n["title"]))
        y -= _LINE * 1.5
        pdf.setFont(default_font, 10)
        dr = i18n["date_range"].format(
            f=date_from.strftime("%d/%m/%Y"), t=date_to.strftime("%d/%m/%Y")
        )
        pdf.drawCentredString(_A4_WIDTH / 2, y, safe(f"{store_name} | {dr}"))
        y -= _LINE * 2

    def _draw_footer() -> None:
        pdf.setFont(default_font, 8)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        generated = safe(f"{i18n['generated']} {now_str}")
        if is_arabic:
            pdf.drawRightString(_A4_WIDTH - _MARGIN, 10 * mm, generated)
            pdf.drawString(_MARGIN, 10 * mm, f"Page {page_num}")
        else:
            pdf.drawString(_MARGIN, 10 * mm, generated)
            pdf.drawRightString(_A4_WIDTH - _MARGIN, 10 * mm, f"Page {page_num}")

    def section_header(title: str) -> None:
        nonlocal y
        new_page_if_needed(3 * _LINE)
        y -= _LINE
        pdf.setFont(bold_font, 14)
        draw_str(_MARGIN, y, safe(title))
        y -= _LINE * 0.5
        pdf.line(_MARGIN, y, _A4_WIDTH - _MARGIN, y)
        y -= _LINE

    def table_row(cols: list[tuple[str, float]], bold: bool = False) -> None:
        nonlocal y
        new_page_if_needed(_LINE)
        pdf.setFont(bold_font if bold else default_font, 10)
        for text, x in cols:
            draw_str(x, y, safe(str(text)))
        y -= _LINE

    def kpi_row(label: str, value: object) -> None:
        nonlocal y
        if is_arabic:
            pdf.drawRightString(_A4_WIDTH - _MARGIN, y, safe(label))
            pdf.drawString(_MARGIN, y, str(value))
        else:
            pdf.drawString(_MARGIN, y, safe(label))
            pdf.drawRightString(_MARGIN + 60 * mm, y, str(value))
        y -= _LINE

    # Initialize first page
    _draw_header()

    # 1. Financial summary
    section_header(i18n["sec_financial"])
    pdf.setFont(default_font, 10)
    kpi_row(i18n["kpi_revenue"], summary.revenue)
    kpi_row(i18n["kpi_gross_profit"], summary.gross_profit)
    kpi_row(i18n["kpi_sales_count"], summary.sales_count)
    kpi_row(i18n["kpi_discounts"], summary.total_discounts)
    kpi_row(i18n["kpi_stock_cost"], inventory.stock_value_cost)
    kpi_row(i18n["kpi_stock_retail"], inventory.stock_value_retail)
    kpi_row(i18n["kpi_supplier_debt"], financial.supplier_debt_total)

    # 2. Top products
    if top_prods:
        section_header(i18n["sec_top_products"])
        col_x = [
            (_MARGIN, i18n["col_product"]),
            (_MARGIN + 80 * mm, i18n["col_quantity"]),
            (_MARGIN + 120 * mm, i18n["col_revenue"]),
        ]
        table_row([(c[1], c[0]) for c in col_x], bold=True)
        for p in top_prods:
            name = p.name[:35] + "…" if len(p.name) > 35 else p.name
            table_row(
                [
                    (name, col_x[0][0]),
                    (p.quantity_sold, col_x[1][0]),
                    (p.revenue, col_x[2][0]),
                ]
            )

    # 3. Top customers
    if top_custs:
        section_header(i18n["sec_top_customers"])
        col_x = [
            (_MARGIN, i18n["col_customer"]),
            (_MARGIN + 60 * mm, i18n["col_phone"]),
            (_MARGIN + 100 * mm, i18n["col_revenue"]),
            (_MARGIN + 140 * mm, i18n["col_sales"]),
        ]
        table_row([(c[1], c[0]) for c in col_x], bold=True)
        for c in top_custs:
            name = c.name[:25] + "…" if len(c.name) > 25 else c.name
            table_row(
                [
                    (name, col_x[0][0]),
                    (c.phone, col_x[1][0]),
                    (c.revenue, col_x[2][0]),
                    (c.sales_count, col_x[3][0]),
                ]
            )

    # 3b. Sales by category
    if category_stats:
        section_header(i18n["sec_category"])
        col_x = [
            (_MARGIN, i18n["col_category"]),
            (_MARGIN + 70 * mm, i18n["col_revenue"]),
            (_MARGIN + 105 * mm, i18n["col_profit"]),
            (_MARGIN + 140 * mm, i18n["col_margin"]),
        ]
        table_row([(c[1], c[0]) for c in col_x], bold=True)
        for cat in category_stats:
            cname = cat.name or i18n["no_category"]
            cname = cname[:30] + "…" if len(cname) > 30 else cname
            margin = (cat.profit / cat.revenue * 100) if cat.revenue else Decimal("0")
            table_row(
                [
                    (cname, col_x[0][0]),
                    (cat.revenue, col_x[1][0]),
                    (cat.profit, col_x[2][0]),
                    (f"{margin:.0f} %", col_x[3][0]),
                ]
            )

    # 4. Payment methods
    if payments:
        section_header(i18n["sec_payments"])
        col_x = [
            (_MARGIN, i18n["col_mode"]),
            (_MARGIN + 60 * mm, i18n["col_amount"]),
            (_MARGIN + 120 * mm, i18n["col_transactions"]),
        ]
        table_row([(c[1], c[0]) for c in col_x], bold=True)
        for p in payments:
            table_row(
                [
                    (
                        payment_labels.get(p.payment_method, p.payment_method),
                        col_x[0][0],
                    ),
                    (p.total, col_x[1][0]),
                    (p.count, col_x[2][0]),
                ]
            )

    # 5. Low stock
    if low_stock:
        section_header(i18n["sec_low_stock"])
        col_x = [
            (_MARGIN, i18n["col_product"]),
            (_MARGIN + 80 * mm, i18n["col_stock"]),
            (_MARGIN + 120 * mm, i18n["col_threshold"]),
        ]
        table_row([(c[1], c[0]) for c in col_x], bold=True)
        for p in low_stock:
            name = p.name[:35] + "…" if len(p.name) > 35 else p.name
            table_row(
                [
                    (name, col_x[0][0]),
                    (p.stock_quantity, col_x[1][0]),
                    (p.low_stock_threshold, col_x[2][0]),
                ]
            )

    # 5b. Dormant stock
    if dead:
        section_header(i18n["sec_dead_stock"])
        col_x = [
            (_MARGIN, i18n["col_product"]),
            (_MARGIN + 90 * mm, i18n["col_tied"]),
            (_MARGIN + 130 * mm, i18n["col_days_dormant"]),
        ]
        table_row([(c[1], c[0]) for c in col_x], bold=True)
        for d in dead:
            dname = d.name[:35] + "…" if len(d.name) > 35 else d.name
            days = (
                i18n["never_sold"]
                if d.days_since is None
                else f"{d.days_since} {i18n['day_abbr']}"
            )
            table_row(
                [
                    (dname, col_x[0][0]),
                    (d.tied_capital, col_x[1][0]),
                    (days, col_x[2][0]),
                ]
            )

    # 6. Outstanding credits
    if credits:
        section_header(i18n["sec_credits"])
        col_x = [
            (_MARGIN, i18n["col_customer"]),
            (_MARGIN + 50 * mm, i18n["col_total"]),
            (_MARGIN + 80 * mm, i18n["col_paid"]),
            (_MARGIN + 110 * mm, i18n["col_balance"]),
            (_MARGIN + 140 * mm, i18n["col_age"]),
        ]
        table_row([(c[1], c[0]) for c in col_x], bold=True)
        for c in credits:
            cname = c.customer_name or i18n["unknown"]
            name = cname[:20] + "…" if len(cname) > 20 else cname
            table_row(
                [
                    (name, col_x[0][0]),
                    (c.total_amount, col_x[1][0]),
                    (c.paid_amount, col_x[2][0]),
                    (c.balance, col_x[3][0]),
                    (f"{c.age_days} {i18n['day_abbr']}", col_x[4][0]),
                ]
            )

    _draw_footer()
    pdf.save()
    return buffer.getvalue()


def build_summary_report_xlsx(
    db: Session,
    store_id: UUID,
    date_from: datetime,
    date_to: datetime,
    language: str = "fr",
) -> bytes:
    """Multi-sheet Excel export of the same data as the PDF."""
    i18n = _REPORT_I18N.get(language, _REPORT_I18N["fr"])
    payment_labels = {
        "cash": i18n["payment_cash"],
        "card": i18n["payment_card"],
        "mobile": i18n["payment_mobile"],
        "other": i18n["payment_other"],
    }

    summary = statistics.sales_summary(db, store_id, date_from, date_to)
    top_prods = statistics.top_products(db, store_id, date_from, date_to, limit=10)
    top_custs = customers.top_customers(db, store_id, date_from, date_to, limit=10)
    category_stats = statistics.category_breakdown(db, store_id, date_from, date_to)
    payments = statistics.payment_method_breakdown(db, store_id, date_from, date_to)
    inventory = statistics.inventory_stats(db, store_id)
    financial = statistics.financial_snapshot(db, store_id)
    low_stock = alerts.low_stock_products(db, store_id)
    dead = statistics.dead_stock(db, store_id, days=60, limit=15)
    credits = alerts.outstanding_credits(db, store_id)

    wb = Workbook()

    def format_headers(ws, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    def write_decimal(ws, row: list) -> None:
        new_row = []
        for val in row:
            if isinstance(val, Decimal):
                new_row.append(float(val))
            else:
                new_row.append(_sanitize_cell(val))
        ws.append(new_row)
        for cell in ws[ws.max_row]:
            if isinstance(cell.value, float):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # 1. Summary
    ws_syn = wb.active
    ws_syn.title = i18n["sheet_summary"]
    format_headers(ws_syn, [i18n["xlsx_metric"], i18n["xlsx_value"]])
    write_decimal(ws_syn, [i18n["xlsx_revenue"], summary.revenue])
    write_decimal(ws_syn, [i18n["xlsx_gross_profit"], summary.gross_profit])
    write_decimal(ws_syn, [i18n["xlsx_num_sales"], summary.sales_count])
    write_decimal(ws_syn, [i18n["xlsx_discounts"], summary.total_discounts])
    write_decimal(ws_syn, [i18n["xlsx_stock_cost"], inventory.stock_value_cost])
    write_decimal(ws_syn, [i18n["xlsx_stock_retail"], inventory.stock_value_retail])
    write_decimal(ws_syn, [i18n["xlsx_supplier_debt"], financial.supplier_debt_total])
    ws_syn.column_dimensions["A"].width = 25
    ws_syn.column_dimensions["B"].width = 15

    # 2. Top products
    ws_prod = wb.create_sheet(i18n["sheet_top_products"])
    format_headers(
        ws_prod,
        [
            i18n["col_product"],
            i18n["xlsx_qty_sold"],
            i18n["xlsx_revenue"],
            i18n["xlsx_profit"],
        ],
    )
    for p in top_prods:
        write_decimal(ws_prod, [p.name, p.quantity_sold, p.revenue, p.profit])
    ws_prod.column_dimensions["A"].width = 30
    ws_prod.column_dimensions["B"].width = 20
    ws_prod.column_dimensions["C"].width = 20
    ws_prod.column_dimensions["D"].width = 20

    # 3. Top customers
    ws_cust = wb.create_sheet(i18n["sheet_top_customers"])
    format_headers(
        ws_cust,
        [
            i18n["col_customer"],
            i18n["col_phone"],
            i18n["xlsx_revenue"],
            i18n["xlsx_num_sales"],
        ],
    )
    for c in top_custs:
        write_decimal(ws_cust, [c.name, c.phone, c.revenue, c.sales_count])
    ws_cust.column_dimensions["A"].width = 30
    ws_cust.column_dimensions["B"].width = 15
    ws_cust.column_dimensions["C"].width = 15
    ws_cust.column_dimensions["D"].width = 20

    # 3b. Sales by category
    ws_cat = wb.create_sheet(i18n["sheet_category"])
    format_headers(
        ws_cat,
        [
            i18n["col_category"],
            i18n["xlsx_revenue"],
            i18n["xlsx_profit"],
            i18n["xlsx_quantity"],
        ],
    )
    for cat in category_stats:
        write_decimal(
            ws_cat,
            [cat.name or i18n["no_category"], cat.revenue, cat.profit, cat.quantity],
        )
    ws_cat.column_dimensions["A"].width = 30
    ws_cat.column_dimensions["B"].width = 18
    ws_cat.column_dimensions["C"].width = 18
    ws_cat.column_dimensions["D"].width = 15

    # 4. Payment methods
    ws_pay = wb.create_sheet(i18n["sheet_payments"])
    format_headers(
        ws_pay,
        [i18n["col_mode"], i18n["col_amount"], i18n["col_transactions"]],
    )
    for p in payments:
        write_decimal(
            ws_pay,
            [payment_labels.get(p.payment_method, p.payment_method), p.total, p.count],
        )
    ws_pay.column_dimensions["A"].width = 20
    ws_pay.column_dimensions["B"].width = 15
    ws_pay.column_dimensions["C"].width = 15

    # 5. Low stock
    ws_stock = wb.create_sheet(i18n["sheet_low_stock"])
    format_headers(
        ws_stock,
        [i18n["col_product"], i18n["xlsx_current_stock"], i18n["xlsx_alert_threshold"]],
    )
    for p in low_stock:
        write_decimal(ws_stock, [p.name, p.stock_quantity, p.low_stock_threshold])
    ws_stock.column_dimensions["A"].width = 30
    ws_stock.column_dimensions["B"].width = 15
    ws_stock.column_dimensions["C"].width = 15

    # 5b. Dormant stock
    ws_dead = wb.create_sheet(i18n["sheet_dead_stock"])
    format_headers(
        ws_dead,
        [i18n["col_product"], i18n["xlsx_tied"], i18n["xlsx_days_dormant"]],
    )
    for d in dead:
        days = i18n["never_sold"] if d.days_since is None else d.days_since
        write_decimal(ws_dead, [d.name, d.tied_capital, days])
    ws_dead.column_dimensions["A"].width = 30
    ws_dead.column_dimensions["B"].width = 18
    ws_dead.column_dimensions["C"].width = 18

    # 6. Outstanding credits
    ws_cred = wb.create_sheet(i18n["sheet_credits"])
    format_headers(
        ws_cred,
        [
            i18n["col_customer"],
            i18n["col_total"],
            i18n["col_paid"],
            i18n["col_balance"],
            i18n["xlsx_age_days"],
        ],
    )
    for c in credits:
        write_decimal(
            ws_cred,
            [c.customer_name, c.total_amount, c.paid_amount, c.balance, c.age_days],
        )
    ws_cred.column_dimensions["A"].width = 30
    ws_cred.column_dimensions["B"].width = 15
    ws_cred.column_dimensions["C"].width = 15
    ws_cred.column_dimensions["D"].width = 15
    ws_cred.column_dimensions["E"].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _open_report(language: str):
    """Shared A4 report scaffolding: (pdf, buffer, state, helpers).

    Returns the canvas, its buffer, a small mutable state dict (y, page) and a
    bundle of drawing closures (safe, draw_str, section_header, table_row,
    kpi_row, new_page_if_needed, draw_footer) — the same visual language as
    build_summary_report_pdf, factored so the debt/daily reports stay DRY.
    """
    is_arabic = language == "ar"
    i18n = _REPORT_I18N.get(language, _REPORT_I18N["fr"])
    default_font = "Amiri" if is_arabic else "Helvetica"
    bold_font = "Amiri" if is_arabic else "Helvetica-Bold"

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    state = {"y": _A4_HEIGHT - _MARGIN, "page": 1}

    def safe(text: object) -> str:
        return _safe(str(text), is_arabic)

    def draw_str(x: float, y_pos: float, text: str) -> None:
        if is_arabic:
            pdf.drawRightString(_A4_WIDTH - x, y_pos, text)
        else:
            pdf.drawString(x, y_pos, text)

    def draw_footer() -> None:
        pdf.setFont(default_font, 8)
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        generated = safe(f"{i18n['generated']} {now_str}")
        if is_arabic:
            pdf.drawRightString(_A4_WIDTH - _MARGIN, 10 * mm, generated)
            pdf.drawString(_MARGIN, 10 * mm, f"Page {state['page']}")
        else:
            pdf.drawString(_MARGIN, 10 * mm, generated)
            pdf.drawRightString(_A4_WIDTH - _MARGIN, 10 * mm, f"Page {state['page']}")

    def new_page_if_needed(needed: float) -> None:
        if state["y"] - needed < _MARGIN + 12 * mm:
            draw_footer()
            pdf.showPage()
            state["y"] = _A4_HEIGHT - _MARGIN
            state["page"] += 1

    def section_header(title: str) -> None:
        new_page_if_needed(3 * _LINE)
        state["y"] -= _LINE
        pdf.setFont(bold_font, 14)
        draw_str(_MARGIN, state["y"], safe(title))
        state["y"] -= _LINE * 0.5
        pdf.line(_MARGIN, state["y"], _A4_WIDTH - _MARGIN, state["y"])
        state["y"] -= _LINE

    def table_row(cols: list[tuple[str, float]], bold: bool = False) -> None:
        new_page_if_needed(_LINE)
        pdf.setFont(bold_font if bold else default_font, 10)
        for text, x in cols:
            draw_str(x, state["y"], safe(text))
        state["y"] -= _LINE

    def kpi_row(label: str, value: object) -> None:
        new_page_if_needed(_LINE)
        pdf.setFont(default_font, 11)
        if is_arabic:
            pdf.drawRightString(_A4_WIDTH - _MARGIN, state["y"], safe(label))
            pdf.drawString(_MARGIN, state["y"], safe(value))
        else:
            pdf.drawString(_MARGIN, state["y"], safe(label))
            pdf.drawRightString(_MARGIN + 70 * mm, state["y"], safe(value))
        state["y"] -= _LINE

    helpers = {
        "safe": safe,
        "draw_str": draw_str,
        "draw_footer": draw_footer,
        "new_page_if_needed": new_page_if_needed,
        "section_header": section_header,
        "table_row": table_row,
        "kpi_row": kpi_row,
        "default_font": default_font,
        "bold_font": bold_font,
        "is_arabic": is_arabic,
        "i18n": i18n,
    }
    return pdf, buffer, state, helpers


def build_debt_report_pdf(db: Session, store_id: UUID, language: str = "fr") -> bytes:
    """A4 summary of every outstanding customer debt (Créances export)."""
    store = db.scalar(select(Store).where(Store.id == store_id))
    store_name = store.name if store else "Ma Boutique"

    credits = alerts.outstanding_credits(db, store_id)
    total = sum((c.balance for c in credits), Decimal("0.00"))

    pdf, buffer, state, h = _open_report(language)
    i18n = h["i18n"]

    pdf.setFont(h["bold_font"], 18)
    pdf.drawCentredString(_A4_WIDTH / 2, state["y"], h["safe"](i18n["debt_title"]))
    state["y"] -= _LINE * 1.5
    pdf.setFont(h["default_font"], 10)
    pdf.drawCentredString(_A4_WIDTH / 2, state["y"], h["safe"](store_name))
    state["y"] -= _LINE * 2

    h["kpi_row"](i18n["debt_total_label"], _fmt_money(total))
    h["kpi_row"](i18n["debtors_label"], str(len(credits)))

    if credits:
        h["section_header"](i18n["sec_credits"])
        col_x = [
            (i18n["col_customer"], _MARGIN),
            (i18n["col_phone"], _MARGIN + 50 * mm),
            (i18n["col_total"], _MARGIN + 88 * mm),
            (i18n["col_paid"], _MARGIN + 112 * mm),
            (i18n["col_balance"], _MARGIN + 136 * mm),
            (i18n["col_age"], _MARGIN + 160 * mm),
        ]
        h["table_row"](col_x, bold=True)
        for c in credits:
            cname = c.customer_name or i18n["anonymous"]
            cname = cname[:22] + "…" if len(cname) > 22 else cname
            h["table_row"](
                [
                    (cname, _MARGIN),
                    (c.customer_phone or "—", _MARGIN + 50 * mm),
                    (_fmt_money(c.total_amount), _MARGIN + 88 * mm),
                    (_fmt_money(c.paid_amount), _MARGIN + 112 * mm),
                    (_fmt_money(c.balance), _MARGIN + 136 * mm),
                    (f"{c.age_days} {i18n['day_abbr']}", _MARGIN + 160 * mm),
                ]
            )

    h["draw_footer"]()
    pdf.save()
    return buffer.getvalue()


def build_daily_report_pdf(
    db: Session,
    store_id: UUID,
    day: date,
    store: Store,
    store_settings: StoreSettings | None,
) -> bytes:
    """End-of-day A4 report for one calendar day (any past date)."""
    language = getattr(store_settings, "ui_language", "fr") if store_settings else "fr"
    store_name = (store_settings.shop_name if store_settings else None) or (
        store.name if store else "Ma Boutique"
    )
    address = store_settings.address if store_settings else None

    start = statistics.to_utc_naive(datetime.combine(day, time.min))
    end = statistics.to_utc_naive(datetime.combine(day, time.max))

    summary = day_closing.day_summary(db, store_id, day)
    top_prods = statistics.top_products(db, store_id, start, end, limit=10)
    day_sales = sales.list_sales(db, store_id, date_from=start, date_to=end)

    net_revenue = summary.total_revenue - summary.total_refunds
    avg_basket = (
        (summary.total_revenue / summary.sales_count)
        if summary.sales_count
        else Decimal("0.00")
    )

    pdf, buffer, state, h = _open_report(language)
    i18n = h["i18n"]
    payment_labels = {
        "cash": i18n["payment_cash"],
        "card": i18n["payment_card"],
        "transfer": i18n["payment_transfer"],
        "mobile": i18n["payment_mobile"],
        "other": i18n["payment_other"],
    }

    pdf.setFont(h["bold_font"], 18)
    pdf.drawCentredString(_A4_WIDTH / 2, state["y"], h["safe"](i18n["daily_title"]))
    state["y"] -= _LINE * 1.5
    pdf.setFont(h["default_font"], 10)
    header_line = f"{store_name} | {day.strftime('%d/%m/%Y')}"
    pdf.drawCentredString(_A4_WIDTH / 2, state["y"], h["safe"](header_line))
    state["y"] -= _LINE
    if address:
        pdf.drawCentredString(_A4_WIDTH / 2, state["y"], h["safe"](address))
        state["y"] -= _LINE
    state["y"] -= _LINE

    # Section 1 — sales summary.
    h["section_header"](i18n["daily_sec_summary"])
    h["kpi_row"](i18n["kpi_transactions"], str(summary.sales_count))
    h["kpi_row"](i18n["kpi_revenue"], _fmt_money(summary.total_revenue))
    h["kpi_row"](i18n["kpi_discounts"], _fmt_money(summary.total_discounts))
    h["kpi_row"](i18n["kpi_refunds"], _fmt_money(summary.total_refunds))
    h["kpi_row"](i18n["kpi_net_revenue"], _fmt_money(net_revenue))
    h["kpi_row"](i18n["kpi_avg_basket"], _fmt_money(avg_basket))

    # Section 2 — payment methods.
    h["section_header"](i18n["daily_sec_payments"])
    method_totals = [
        ("cash", summary.cash_total),
        ("card", summary.card_total),
        ("transfer", summary.transfer_total),
        ("other", summary.other_total),
    ]
    h["table_row"](
        [
            (i18n["col_mode"], _MARGIN),
            (i18n["col_amount"], _MARGIN + 90 * mm),
        ],
        bold=True,
    )
    for method, amount in method_totals:
        if amount > 0:
            h["table_row"](
                [
                    (payment_labels.get(method, method), _MARGIN),
                    (_fmt_money(amount), _MARGIN + 90 * mm),
                ]
            )

    # Section 3 — top 10 products.
    if top_prods:
        h["section_header"](i18n["daily_sec_top"])
        h["table_row"](
            [
                (i18n["col_product"], _MARGIN),
                (i18n["col_quantity"], _MARGIN + 100 * mm),
                (i18n["col_revenue"], _MARGIN + 135 * mm),
            ],
            bold=True,
        )
        for p in top_prods:
            name = p.name[:40] + "…" if len(p.name) > 40 else p.name
            h["table_row"](
                [
                    (name, _MARGIN),
                    (str(p.quantity_sold), _MARGIN + 100 * mm),
                    (_fmt_money(p.revenue), _MARGIN + 135 * mm),
                ]
            )

    # Section 4 — detailed sales log.
    if day_sales:
        h["section_header"](i18n["daily_sec_log"])
        h["table_row"](
            [
                (i18n["col_time"], _MARGIN),
                (i18n["col_invoice"], _MARGIN + 22 * mm),
                (i18n["col_customer"], _MARGIN + 52 * mm),
                (i18n["col_items"], _MARGIN + 105 * mm),
                (i18n["col_total"], _MARGIN + 128 * mm),
                (i18n["col_payment"], _MARGIN + 158 * mm),
            ],
            bold=True,
        )
        for sale in day_sales:
            created = sale.created_at
            time_str = created.strftime("%H:%M") if created else "—"
            invoice = (
                f"{sale.invoice_number:06d}"
                if sale.invoice_number
                else str(sale.id)[:6].upper()
            )
            cname = getattr(sale, "customer_name", None) or i18n["anonymous"]
            cname = cname[:24] + "…" if len(cname) > 24 else cname
            live_items = [it for it in sale.items if it.deleted_at is None]
            item_count = sum(it.quantity for it in live_items)
            method = "—"
            live_payments = [p for p in sale.payments if p.deleted_at is None]
            if live_payments:
                method = payment_labels.get(
                    live_payments[0].payment_method, live_payments[0].payment_method
                )
            h["table_row"](
                [
                    (time_str, _MARGIN),
                    (invoice, _MARGIN + 22 * mm),
                    (cname, _MARGIN + 52 * mm),
                    (str(item_count), _MARGIN + 105 * mm),
                    (_fmt_money(sale.total_amount), _MARGIN + 128 * mm),
                    (method, _MARGIN + 158 * mm),
                ]
            )

    # Footer signature line.
    state["y"] -= _LINE
    h["new_page_if_needed"](2 * _LINE)
    pdf.setFont(h["default_font"], 10)
    h["draw_str"](
        _MARGIN, state["y"], h["safe"](f"{i18n['signature_manager']} : ____________")
    )
    state["y"] -= _LINE

    h["draw_footer"]()
    pdf.save()
    return buffer.getvalue()


def build_comparison_report_pdf(
    db: Session,
    store_id: UUID,
    a_from: datetime,
    a_to: datetime,
    b_from: datetime,
    b_to: datetime,
    language: str = "fr",
) -> bytes:
    """A4 side-by-side comparison of two periods (metrics + top products)."""
    store = db.scalar(select(Store).where(Store.id == store_id))
    store_name = store.name if store else "Ma Boutique"

    sum_a = statistics.sales_summary(db, store_id, a_from, a_to)
    sum_b = statistics.sales_summary(db, store_id, b_from, b_to)
    top_a = statistics.top_products(db, store_id, a_from, a_to, limit=5)
    top_b = statistics.top_products(db, store_id, b_from, b_to, limit=5)

    def _basket(summary) -> Decimal:
        return (
            (summary.revenue / summary.sales_count)
            if summary.sales_count
            else Decimal("0.00")
        )

    def _change(a: Decimal, b: Decimal) -> str:
        a, b = Decimal(str(a)), Decimal(str(b))
        if b == 0:
            return "-" if a == 0 else "+100 %"
        pct = (a - b) / b * 100
        sign = "+" if pct >= 0 else ""
        return f"{sign}{pct:.0f} %"

    pdf, buffer, state, h = _open_report(language)
    i18n = h["i18n"]

    pdf.setFont(h["bold_font"], 18)
    pdf.drawCentredString(
        _A4_WIDTH / 2, state["y"], h["safe"](i18n["comparison_title"])
    )
    state["y"] -= _LINE * 1.5
    pdf.setFont(h["default_font"], 10)
    dr = (
        f"{store_name}  |  A: {a_from.strftime('%d/%m/%Y')}-{a_to.strftime('%d/%m/%Y')}"
        f"   B: {b_from.strftime('%d/%m/%Y')}-{b_to.strftime('%d/%m/%Y')}"
    )
    pdf.drawCentredString(_A4_WIDTH / 2, state["y"], h["safe"](dr))
    state["y"] -= _LINE * 2

    col = [
        (i18n["xlsx_metric"], _MARGIN),
        (i18n["cmp_period_a"], _MARGIN + 70 * mm),
        (i18n["cmp_period_b"], _MARGIN + 110 * mm),
        (i18n["cmp_change"], _MARGIN + 150 * mm),
    ]
    h["table_row"](col, bold=True)
    rows = [
        (i18n["xlsx_revenue"], sum_a.revenue, sum_b.revenue, True),
        (i18n["xlsx_gross_profit"], sum_a.gross_profit, sum_b.gross_profit, True),
        (i18n["xlsx_num_sales"], sum_a.sales_count, sum_b.sales_count, False),
        (i18n["cmp_basket"], _basket(sum_a), _basket(sum_b), True),
        (i18n["xlsx_discounts"], sum_a.total_discounts, sum_b.total_discounts, True),
    ]
    for label, a_val, b_val, money in rows:
        a_text = _fmt_money(a_val) if money else str(a_val)
        b_text = _fmt_money(b_val) if money else str(b_val)
        h["table_row"](
            [
                (label, _MARGIN),
                (a_text, _MARGIN + 70 * mm),
                (b_text, _MARGIN + 110 * mm),
                (
                    _change(Decimal(str(a_val)), Decimal(str(b_val))),
                    _MARGIN + 150 * mm,
                ),
            ]
        )

    def _top_section(title: str, top) -> None:
        h["section_header"](title)
        h["table_row"](
            [
                (i18n["col_product"], _MARGIN),
                (i18n["col_quantity"], _MARGIN + 110 * mm),
                (i18n["col_revenue"], _MARGIN + 145 * mm),
            ],
            bold=True,
        )
        for p in top:
            name = p.name[:40] + "…" if len(p.name) > 40 else p.name
            h["table_row"](
                [
                    (name, _MARGIN),
                    (str(p.quantity_sold), _MARGIN + 110 * mm),
                    (_fmt_money(p.revenue), _MARGIN + 145 * mm),
                ]
            )

    if top_a:
        _top_section(f"{i18n['cmp_period_a']} - {i18n['sec_top_products']}", top_a)
    if top_b:
        _top_section(f"{i18n['cmp_period_b']} - {i18n['sec_top_products']}", top_b)

    h["draw_footer"]()
    pdf.save()
    return buffer.getvalue()
