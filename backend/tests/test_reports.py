import io
from datetime import datetime
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.schemas.product import ProductCreate
from app.schemas.sale import CartItem, CheckoutRequest
from app.schemas.store import StoreCreate
from app.services import products, reports, sales, statistics, stores

JAN_15 = datetime(2026, 1, 15, 12, 0)
JAN_RANGE = (datetime(2026, 1, 1), datetime(2026, 1, 31, 23, 59, 59))


def make_product(db, store, name, cost, floor, stock=1000):
    return products.create_product(
        db,
        ProductCreate(
            store_id=store.id,
            name=name,
            cost_price=Decimal(cost),
            price_detail=Decimal("999.00"),
            price_gros=Decimal("999.00"),
            price_super_gros=Decimal(floor),
            stock_quantity=stock,
        ),
    )


def sell(db, store, product, quantity, price, created_at):
    sale = sales.finalize_sale(
        db,
        CheckoutRequest(
            store_id=store.id,
            items=[
                CartItem(
                    product_id=product.id,
                    quantity=quantity,
                    unit_price_override=Decimal(price),
                )
            ],
        ),
    )
    sale.created_at = created_at
    db.commit()
    return sale


def test_pdf_is_valid_for_populated_range(db):
    store = stores.create_store(db, StoreCreate(name="Test"))
    p1 = make_product(db, store, "A", "33.33", "30.00")
    sell(db, store, p1, 2, "49.99", JAN_15)

    pdf_bytes = reports.build_summary_report_pdf(
        db, store.id, JAN_RANGE[0], JAN_RANGE[1]
    )
    assert pdf_bytes.startswith(b"%PDF")
    assert len(pdf_bytes) > 500


def test_pdf_is_valid_for_empty_range(db):
    store = stores.create_store(db, StoreCreate(name="Test"))
    pdf_bytes = reports.build_summary_report_pdf(
        db, store.id, JAN_RANGE[0], JAN_RANGE[1]
    )
    assert pdf_bytes.startswith(b"%PDF")
    assert len(pdf_bytes) > 100


def test_xlsx_values_match_summary_exactly(db):
    store = stores.create_store(db, StoreCreate(name="Test"))
    p1 = make_product(db, store, "A", "10.00", "5.00")
    sell(db, store, p1, 1, "20.00", JAN_15)

    summary = statistics.sales_summary(db, store.id, JAN_RANGE[0], JAN_RANGE[1])
    xlsx_bytes = reports.build_summary_report_xlsx(
        db, store.id, JAN_RANGE[0], JAN_RANGE[1]
    )

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws_syn = wb["Synthèse"]

    # Check values in rows 2, 3, 4, 5 against summary
    assert ws_syn["A2"].value == "Chiffre d'affaires"
    assert ws_syn["B2"].value == float(summary.revenue)

    assert ws_syn["A3"].value == "Marge brute"
    assert ws_syn["B3"].value == float(summary.gross_profit)


def test_xlsx_has_all_sheets(db):
    store = stores.create_store(db, StoreCreate(name="Test"))
    xlsx_bytes = reports.build_summary_report_xlsx(
        db, store.id, JAN_RANGE[0], JAN_RANGE[1]
    )
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == [
        "Synthèse",
        "Meilleures ventes",
        "Meilleurs clients",
        "Modes de paiement",
        "Stock faible",
        "Crédits en attente",
    ]


def test_xlsx_empty_range_has_zero_values(db):
    store = stores.create_store(db, StoreCreate(name="Test"))
    xlsx_bytes = reports.build_summary_report_xlsx(
        db, store.id, JAN_RANGE[0], JAN_RANGE[1]
    )
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws_syn = wb["Synthèse"]
    assert ws_syn["B2"].value == 0.0


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@", "\t", "\r"])
def test_sanitize_cell_prefixes_dangerous_strings(prefix):
    dangerous = f"{prefix}cmd|'/c calc'!A1"
    out = reports._sanitize_cell(dangerous)
    assert out == "'" + dangerous
    assert out.startswith("'")


def test_sanitize_cell_leaves_safe_values_untouched():
    assert reports._sanitize_cell("Eau minérale") == "Eau minérale"
    assert reports._sanitize_cell("0612345678") == "0612345678"
    assert reports._sanitize_cell(42) == 42
    assert reports._sanitize_cell(None) is None


def test_xlsx_sanitizes_formula_injection_in_product_names(db):
    """Product names beginning with a dangerous char land as literal text."""
    store = stores.create_store(db, StoreCreate(name="Test"))
    dangerous_names = ["=DANGER", "+DANGER", "-DANGER", "@DANGER"]
    for name in dangerous_names:
        p = make_product(db, store, name, "10.00", "5.00")
        sell(db, store, p, 1, "20.00", JAN_15)

    xlsx_bytes = reports.build_summary_report_xlsx(
        db, store.id, JAN_RANGE[0], JAN_RANGE[1]
    )
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Meilleures ventes"]
    names_in_sheet = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    for name in dangerous_names:
        assert ("'" + name) in names_in_sheet, f"{name!r} not sanitized"
        assert name not in names_in_sheet, f"{name!r} written raw (formula risk)"
